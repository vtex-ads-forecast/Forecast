#!/usr/bin/env python3
"""
VTEX Ads — Export historical data to Excel.
Pulls Jan-May 2026 from Metabase at advertiser level with AdTech/AdNetwork split.

Usage:
  python scripts/export_excel.py
  python scripts/export_excel.py 2026-01-01 2026-05-31

Environment variables:
  METABASE_USER  — Metabase login email
  METABASE_PASS  — Metabase login password
"""

import os, sys, json, re, requests
from datetime import datetime
from collections import defaultdict

METABASE_URL = "https://metabase.newtail.com.br"
DB_ID = 13
HTML_PATH = os.path.join(os.path.dirname(__file__), "..", "index.html")
SETTINGS_PATH = os.path.join(os.path.dirname(__file__), "..", "settings.json")

FX_RATES = {"BRL": 1.0, "ARS": 0.0036, "COP": 0.0015, "PEN": 1.50}
EXCLUDE_PATTERNS = ["teste", "test", "staging", "hml", "homolog"]


def metabase_auth():
    user = os.environ.get("METABASE_USER")
    pwd = os.environ.get("METABASE_PASS")
    if not user or not pwd:
        raise ValueError("METABASE_USER and METABASE_PASS required")
    resp = requests.post(f"{METABASE_URL}/api/session",
                         json={"username": user, "password": pwd}, timeout=30)
    resp.raise_for_status()
    print(f"✓ Authenticated as {user}")
    return resp.json()["id"]


def fetch_data(token, start_date, end_date):
    headers = {"X-Metabase-Session": token}
    PAGE_SIZE = 2000
    base_sql = f"""
    with metrics AS (
        SELECT cmnd.day, cmnd.campaign_id, cmnd.publisher_id, cmnd.advertiser_id,
               SUM(cmnd.total_clicks_cost) + SUM(cmnd.total_impressions_cost) AS total_cost,
               p.name AS publisher_name, p.currency_code, p.is_test
        FROM CAMPAIGNS_METRICS_NETWORK_DAY cmnd
        JOIN publishers p ON p.id = cmnd.publisher_id
        WHERE day BETWEEN ('{start_date}')::timestamp - INTERVAL '1 day'
                      AND ('{end_date}')::timestamp + INTERVAL '2 days'
        GROUP BY cmnd.day, cmnd.campaign_id, cmnd.publisher_id, cmnd.advertiser_id,
                 p.name, p.currency_code, p.is_test
    ),
    costs AS (
        SELECT r.day, r.publisher_id, r.publisher_name, r.advertiser_id,
               a.name AS advertiser_name, r.currency_code, r.campaign_id,
               SUM(r.total_cost) AS total_cost
        FROM metrics r
        LEFT JOIN advertisers a ON a.id = r.advertiser_id
        WHERE r.day >= '{start_date}'
          AND r.day < '{end_date}'::date + INTERVAL '1 day'
          AND r.is_test = false
        GROUP BY r.day, r.publisher_id, r.publisher_name, r.advertiser_id,
                 a.name, r.currency_code, r.campaign_id
    )
    SELECT day, publisher_id, publisher_name, advertiser_id, advertiser_name,
           currency_code, campaign_id, total_cost
    FROM costs WHERE total_cost > 0 ORDER BY day DESC, publisher_id, advertiser_id
    """

    all_rows, cols = [], []
    offset = 0
    while True:
        sql = f"{base_sql} LIMIT {PAGE_SIZE} OFFSET {offset}"
        print(f"  Fetching offset={offset}...")
        resp = requests.post(f"{METABASE_URL}/api/dataset", headers=headers,
                             json={"database": DB_ID, "type": "native",
                                   "native": {"query": sql}}, timeout=300)
        resp.raise_for_status()
        result = resp.json()
        if not cols:
            cols = [c.get("name", f"col_{i}") for i, c in enumerate(result["data"]["cols"])]
        page = result["data"]["rows"]
        print(f"  Got {len(page)} rows")
        all_rows.extend(page)
        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
        if offset > 200000:
            print("  ⚠ Safety limit")
            break

    col_idx = {}
    for i, c in enumerate(cols):
        cl = c.lower().replace(" ", "_")
        if "day" in cl or "date" in cl: col_idx["day"] = i
        elif "publisher" in cl and "name" in cl: col_idx["publisher_name"] = i
        elif "advertiser" in cl and "name" in cl: col_idx["advertiser_name"] = i
        elif "currency" in cl: col_idx["currency_code"] = i
        elif "cost" in cl or "total" in cl: col_idx["total_cost"] = i

    data = []
    for row in all_rows:
        data.append({
            "day": str(row[col_idx.get("day", 0)] or "")[:10],
            "publisher_name": str(row[col_idx.get("publisher_name", 2)] or ""),
            "advertiser_name": str(row[col_idx.get("advertiser_name", 4)] or ""),
            "currency_code": str(row[col_idx.get("currency_code", 5)] or "BRL"),
            "total_cost": float(row[col_idx.get("total_cost", 7)] or 0),
        })
    print(f"✓ Total: {len(data)} rows ({start_date} → {end_date})")
    return data


def load_pub_mapping():
    """Load publisher → segment and TR mappings from settings.json + HTML."""
    pub_seg, pub_tr = {}, {}

    if os.path.exists(SETTINGS_PATH):
        with open(SETTINGS_PATH) as f:
            settings = json.load(f)
        for pname, ov in settings.items():
            if ov.get("seg"):
                pub_seg[pname] = ov["seg"]
            pub_tr[pname] = {"tech": ov.get("trTech", 0.1), "net": ov.get("trNetwork", 0.15)}
        print(f"✓ Loaded {len(settings)} publishers from settings.json")

    # Also extract from REAL_APRIL in HTML
    with open(HTML_PATH) as f:
        html = f.read()
    for pm in re.finditer(
        r'"([^"]+)":\{spendReal:\d+,revReal:\d+,spendTech:\d+,spendNetwork:\d+,revTech:\d+,revNetwork:\d+,trTech:([\d.]+),trNetwork:([\d.]+)\}',
        html
    ):
        pname = pm.group(1)
        if pname not in pub_seg:
            # Find segment by looking backwards for the segment header
            seg_m = list(re.finditer(r'"([^"]+)":\{spendReal:', html[:pm.start()]))
            if seg_m:
                pub_seg[pname] = seg_m[-1].group(1)
            pub_tr[pname] = {"tech": float(pm.group(2)), "net": float(pm.group(3))}

    print(f"  Total mappings: {len(pub_seg)}")
    return pub_seg, pub_tr


def build_excel(raw_data, pub_seg, pub_tr, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Aggregate: month × advertiser × publisher × tipo
    agg = defaultdict(lambda: {"spend": 0.0, "rev": 0.0})
    for r in raw_data:
        adv = (r["advertiser_name"] or "").strip()
        if any(x in adv.lower() for x in EXCLUDE_PATTERNS):
            continue
        cost = r["total_cost"]
        if cost <= 0:
            continue
        pub = (r["publisher_name"] or "").strip()
        curr = r["currency_code"]
        fx = FX_RATES.get(curr, 1.0)
        cost_brl = cost * fx
        day = r["day"][:7]  # YYYY-MM
        is_net = "vtexads" in adv.lower()
        tipo = "AdNetwork" if is_net else "AdTech"
        seg = pub_seg.get(pub, "Long Tail")
        tri = pub_tr.get(pub, {"tech": 0.1, "net": 0.15})
        tr = tri["net"] if is_net else tri["tech"]
        rev = cost_brl * tr

        key = (day, adv, tipo, pub, seg)
        agg[key]["spend"] += cost_brl
        agg[key]["rev"] += rev

    rows = []
    for (mes, adv, tipo, pub, seg), vals in sorted(agg.items()):
        if vals["spend"] < 1:
            continue
        rows.append({
            "mes": mes, "advertiser": adv, "tipo": tipo,
            "publisher": pub, "segment": seg,
            "adspend": round(vals["spend"]), "receita": round(vals["rev"])
        })

    print(f"✓ {len(rows)} rows to write")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Mensais"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="0F3E99")
    hdr_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    num_fmt = '#,##0'
    alt_fill = PatternFill("solid", fgColor="F5F7FA")

    headers = ["Mês", "Advertiser", "Tipo", "Publisher", "Segmento", "Ad Spend", "Receita"]
    widths = [12, 48, 14, 48, 20, 16, 16]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["mes"]).font = data_font
        ws.cell(row=i, column=2, value=row["advertiser"]).font = data_font
        ws.cell(row=i, column=3, value=row["tipo"]).font = data_font
        ws.cell(row=i, column=4, value=row["publisher"]).font = data_font
        ws.cell(row=i, column=5, value=row["segment"]).font = data_font
        c_sp = ws.cell(row=i, column=6, value=row["adspend"])
        c_sp.font = data_font; c_sp.number_format = num_fmt; c_sp.alignment = Alignment(horizontal="right")
        c_rv = ws.cell(row=i, column=7, value=row["receita"])
        c_rv.font = data_font; c_rv.number_format = num_fmt; c_rv.alignment = Alignment(horizontal="right")
        if i % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=i, column=c).fill = alt_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(rows)+1}"

    # Resumo
    ws2 = wb.create_sheet("Resumo Mensal")
    for c, h in enumerate(["Mês", "Tipo", "Ad Spend", "Receita", "# Advertisers"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

    summary = defaultdict(lambda: {"spend": 0, "rev": 0, "advs": set()})
    for r in rows:
        k = (r["mes"], r["tipo"])
        summary[k]["spend"] += r["adspend"]
        summary[k]["rev"] += r["receita"]
        summary[k]["advs"].add(r["advertiser"])

    for i, (k, v) in enumerate(sorted(summary.items()), 2):
        ws2.cell(row=i, column=1, value=k[0]).font = data_font
        ws2.cell(row=i, column=2, value=k[1]).font = data_font
        ws2.cell(row=i, column=3, value=v["spend"]).number_format = num_fmt
        ws2.cell(row=i, column=4, value=v["rev"]).number_format = num_fmt
        ws2.cell(row=i, column=5, value=len(v["advs"]))

    for c, w in zip(range(1,6), [12, 14, 16, 16, 16]):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"✓ Saved to {output_path}")


def main():
    if len(sys.argv) >= 3:
        start, end = sys.argv[1], sys.argv[2]
    else:
        start, end = "2026-01-01", "2026-05-20"

    print(f"\n{'='*60}")
    print(f"  VTEX Ads — Export to Excel")
    print(f"  Range: {start} → {end}")
    print(f"{'='*60}\n")

    token = metabase_auth()
    raw = fetch_data(token, start, end)
    if not raw:
        print("⚠ No data. Exiting.")
        return

    pub_seg, pub_tr = load_pub_mapping()
    output = os.path.join(os.path.dirname(__file__), "..", "VTEX_Ads_Dados_Jan_Mai_2026.xlsx")
    build_excel(raw, pub_seg, pub_tr, output)


if __name__ == "__main__":
    main()
